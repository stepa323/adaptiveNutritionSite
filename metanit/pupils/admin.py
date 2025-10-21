from django.shortcuts import render
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.contrib import admin
from openpyxl.utils import get_column_letter

from .models import Class, Pupil, Dish, DailyMenu, WeeklyBreakfasts


@admin.register(Pupil)
class PupilAdmin(admin.ModelAdmin):
    list_display = ['last_name', 'first_name', 'class_group']
    list_filter = ['class_group']
    search_fields = ['last_name', 'first_name']


@admin.register(Dish)
class DishAdmin(admin.ModelAdmin):
    list_display = ['name']
    search_fields = ['name']


@admin.register(DailyMenu)
class DailyMenuAdmin(admin.ModelAdmin):
    list_display = ['date', 'option_1', 'option_2']
    list_filter = ['date']
    date_hierarchy = 'date'


@admin.register(WeeklyBreakfasts)
class WeeklyBreakfastAdmin(admin.ModelAdmin):
    list_display = ['pupil', 'get_class_group', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday',
                    'week_start_date', 'choices_count']
    list_filter = ['pupil__class_group', 'week_start_date']
    search_fields = ['pupil__first_name', 'pupil__last_name']

    def get_class_group(self, obj):
        return obj.pupil.class_group

    get_class_group.short_description = 'Класс'

    def choices_count(self, obj):
        count = sum([1 for field in [obj.monday, obj.tuesday, obj.wednesday,
                                     obj.thursday, obj.friday] if field])
        return f"{count}/5"

    choices_count.short_description = 'Выбрано'

    def get_urls(self):
        urls = super().get_urls()
        from django.urls import path
        custom_urls = [
            path('statistics/', self.admin_site.admin_view(self.statistics_view),
                 name='pupils_weeklybreakfasts_statistics'),
            path('export-excel/', self.admin_site.admin_view(self.export_to_excel),
                 name='pupils_weeklybreakfasts_export_excel'),
        ]
        return custom_urls + urls

    def statistics_view(self, request):
        """Кастомная страница статистики"""
        from django.db.models import Count
        from datetime import datetime, timedelta

        # Получаем текущую неделю
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())

        # Собираем статистику по классам
        classes_stats = []
        from .models import Class, Pupil, WeeklyBreakfasts

        for class_obj in Class.objects.all():
            # Всего учеников в классе
            total_pupils = Pupil.objects.filter(class_group=class_obj).count()

            # Статистика по дням
            day_stats = {}
            days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']

            for day in days:
                # Выборы по этому дню
                choices = WeeklyBreakfasts.objects.filter(
                    week_start_date=week_start,
                    pupil__class_group=class_obj
                ).exclude(**{f"{day}__isnull": True})

                # Подсчет по блюдам
                dish_counts = choices.values(
                    f'{day}__name'
                ).annotate(
                    count=Count('id')
                ).order_by('-count')

                # Формируем статистику как в Excel
                dish_stats = {}
                for item in dish_counts:
                    dish_stats[item[f'{day}__name']] = item['count']

                day_stats[day] = {
                    'dishes': dish_stats,
                    'not_chosen': total_pupils - choices.count(),
                    'chosen': choices.count()
                }

            classes_stats.append({
                'class': class_obj,
                'total_pupils': total_pupils,
                'day_stats': day_stats
            })

        context = {
            'title': f'Статистика завтраков на неделю с {week_start} по {week_start + timedelta(days=4)}',
            'classes_stats': classes_stats,
            'week_start': week_start,
        }

        return render(request, 'admin/breakfast_statistics.html', context)

    def changelist_view(self, request, extra_context=None):
        """Добавляем ссылку на статистику в список"""
        extra_context = extra_context or {}
        extra_context['show_statistics_link'] = True
        return super().changelist_view(request, extra_context=extra_context)

    def export_to_excel(self, request):
        """Экспорт с листами статистики и листами учеников"""
        from django.db.models import Count
        from .models import Class, Pupil, WeeklyBreakfasts

        # Получаем текущую неделю
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=4)

        # Создаем Excel книгу
        wb = openpyxl.Workbook()

        # Удаляем дефолтный лист
        wb.remove(wb.active)

        # Стили
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center')

        # ЛИСТ СТАТИСТИКИ (как у нас было)
        ws_stats = wb.create_sheet(title="Статистика")

        # Заголовок статистики
        ws_stats.merge_cells('A1:E1')
        ws_stats['A1'] = f'Статистика завтраков с {week_start} по {week_end}'
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats['A1'].alignment = center_align

        row = 3

        for class_obj in Class.objects.all().order_by('name'):
            # Заголовок класса
            ws_stats.merge_cells(f'A{row}:E{row}')
            ws_stats[f'A{row}'] = f'Класс: {class_obj.name}'
            ws_stats[f'A{row}'].font = bold_font
            row += 1

            # Всего учеников
            total_pupils = Pupil.objects.filter(class_group=class_obj).count()
            ws_stats[f'A{row}'] = f'Всего учеников: {total_pupils}'
            row += 1

            # Заголовки таблицы
            headers = ['День недели', 'Вариант 1', 'Вариант 2', 'Не выбрано', 'Выбрали']
            for col, header in enumerate(headers, 1):
                cell = ws_stats.cell(row=row, column=col, value=header)
                cell.font = bold_font
                cell.fill = header_fill

            row += 1

            # Статистика по дням
            days_mapping = {
                'monday': 'Понедельник',
                'tuesday': 'Вторник',
                'wednesday': 'Среда',
                'thursday': 'Четверг',
                'friday': 'Пятница'
            }

            for day_field, day_name in days_mapping.items():
                # Выборы по этому дню
                choices = WeeklyBreakfasts.objects.filter(
                    week_start_date=week_start,
                    pupil__class_group=class_obj
                ).exclude(**{f"{day_field}__isnull": True})

                # Подсчет по блюдам
                dish_counts = choices.values(
                    f'{day_field}__name'
                ).annotate(count=Count('id')).order_by('-count')

                # Формируем данные
                dishes_list = list(dish_counts)
                option1 = f"{dishes_list[0][f'{day_field}__name']}: {dishes_list[0]['count']}" if dishes_list else "-"
                option2 = f"{dishes_list[1][f'{day_field}__name']}: {dishes_list[1]['count']}" if len(
                    dishes_list) > 1 else "-"

                not_chosen = total_pupils - choices.count()
                chosen = choices.count()

                # Записываем строку
                ws_stats.cell(row=row, column=1, value=day_name)
                ws_stats.cell(row=row, column=2, value=option1)
                ws_stats.cell(row=row, column=3, value=option2)
                ws_stats.cell(row=row, column=4, value=not_chosen)
                ws_stats.cell(row=row, column=5, value=chosen)

                row += 1

            # Пустая строка между классами
            row += 2

        # ЛИСТЫ С УЧЕНИКАМИ (добавляем для каждого класса)
        for class_obj in Class.objects.all().order_by('name'):
            ws_pupils = wb.create_sheet(title=f"Ученики {class_obj.name}")

            # Заголовок
            ws_pupils.merge_cells('A1:H1')
            ws_pupils['A1'] = f'Выборы завтраков - {class_obj.name} (неделя {week_start})'
            ws_pupils['A1'].font = Font(bold=True, size=14)
            ws_pupils['A1'].alignment = center_align

            # Заголовки таблицы
            headers = ['Фамилия', 'Имя', 'Класс', 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']
            for col, header in enumerate(headers, 1):
                cell = ws_pupils.cell(row=3, column=col, value=header)
                cell.font = bold_font
                cell.fill = header_fill

            # Данные учеников
            pupils = Pupil.objects.filter(class_group=class_obj).order_by('last_name', 'first_name')
            row = 4

            for pupil in pupils:
                try:
                    breakfast = WeeklyBreakfasts.objects.get(pupil=pupil, week_start_date=week_start)
                except WeeklyBreakfasts.DoesNotExist:
                    breakfast = None

                # Данные ученика
                ws_pupils.cell(row=row, column=1, value=pupil.last_name)
                ws_pupils.cell(row=row, column=2, value=pupil.first_name)
                ws_pupils.cell(row=row, column=3, value=class_obj.name)

                # Выборы по дням
                days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
                day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']

                for col, (day_field, day_name) in enumerate(zip(days, day_names), 4):
                    if breakfast and getattr(breakfast, day_field):
                        ws_pupils.cell(row=row, column=col, value=getattr(breakfast, day_field).name)
                    else:
                        ws_pupils.cell(row=row, column=col, value="Не выбрано")

                row += 1

            # Настраиваем ширину колонок
            column_widths = [15, 15, 10, 30, 30, 30, 30, 30]
            for i, width in enumerate(column_widths, 1):
                ws_pupils.column_dimensions[get_column_letter(i)].width = width

        # Настраиваем ширину колонок для листа статистики
        stats_widths = [15, 30, 30, 12, 12]
        for i, width in enumerate(stats_widths, 1):
            ws_stats.column_dimensions[get_column_letter(i)].width = width

        # Создаем HTTP ответ с файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'завтраки_статистика_{week_start}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response